from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from ..formatters import compact_single_line, format_rows
from ..models import AttachmentReaderConfig


class ExcelAttachmentReader:
    """
    負責在 builder.attachment.readers.excel_reader 中封裝 ExcelAttachmentReader，封裝附件讀取與內容萃取流程，將檔案轉成可推理的證據。
    
    Args:
        config: 控制此流程行為的設定資料。
    
    Returns:
        類別本身不直接回傳值；建立實例後可透過其方法操作狀態與流程。
    
    限制或副作用:
        方法可能更新內部狀態、讀寫檔案、呼叫外部服務或產生日誌，需依使用情境確認。
    """
    def __init__(self, config: AttachmentReaderConfig) -> None:
        """
        負責執行 ExcelAttachmentReader 中的 __init__ 流程，初始化物件所需的設定、依賴與內部狀態，讓後續方法可以沿用同一份執行上下文。
        
        Args:
            config: 控制此流程行為的設定資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 None。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        self.config = config

    def read_xlsx(self, question: str, file_path: Path) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 read_xlsx 流程，讀取表格資料並整理成可控長度的文字或結構化摘要。
        
        Args:
            question: 目前要處理的任務、問題或查詢文字。
            file_path: 要讀取或寫入的檔案或目錄路徑。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        try:
            import openpyxl
        except ImportError:
            return "Excel attachment detected, but openpyxl is not installed."

        workbook = openpyxl.load_workbook(str(file_path), read_only=False, data_only=True)
        sections: list[str] = []
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                strategy = self._classify_excel_question(question)
                max_row = self._safe_int(sheet.max_row)
                max_column = self._safe_int(sheet.max_column)
                row_limit = self._excel_row_limit_for_strategy(strategy)
                rows: list[list[str]] = []
                color_rows: list[list[str]] = []
                colored_cells: list[str] = []

                for idx, row in enumerate(sheet.iter_rows()):
                    if idx >= row_limit:
                        break
                    value_row: list[str] = []
                    color_row: list[str] = []
                    for cell in row:
                        value_row.append("" if cell.value is None else str(cell.value))
                        color = self._get_excel_fill_color(cell)
                        if color:
                            color_row.append(f"{cell.coordinate}={color}")
                            colored_cells.append(
                                f"{cell.coordinate} value={compact_single_line(cell.value)} color={color}"
                            )
                    rows.append(value_row)
                    if color_row:
                        color_rows.append(color_row)

                section_parts = [
                    format_rows(
                        f"Sheet {sheet_name} ({max_row} rows x {max_column} columns)",
                        rows,
                        truncated=max_row > row_limit,
                        max_rows=row_limit,
                    ),
                    f"Excel evidence strategy: {strategy}",
                ]

                summary = self._summarize_excel_rows(rows)
                if summary:
                    section_parts.append(summary)

                question_summary = self._build_question_oriented_excel_summary(question, rows)
                if question_summary:
                    section_parts.append(question_summary)

                color_section = self._build_excel_color_section(
                    strategy=strategy,
                    color_rows=color_rows,
                    colored_cells=colored_cells,
                    max_row=max_row,
                    max_column=max_column,
                )
                if color_section:
                    section_parts.append(color_section)

                sections.append("\n\n".join(section_parts).strip())
        finally:
            workbook.close()
        return "\n\n".join(section for section in sections if section)

    def read_xls(self, question: str, file_path: Path) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 read_xls 流程，讀取表格資料並整理成可控長度的文字或結構化摘要。
        
        Args:
            question: 目前要處理的任務、問題或查詢文字。
            file_path: 要讀取或寫入的檔案或目錄路徑。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        try:
            import pandas as pd
        except ImportError:
            return "XLS attachment detected, but pandas is not installed."

        try:
            sheets = pd.read_excel(file_path, sheet_name=None, dtype=object, engine="xlrd")
        except ImportError:
            return "XLS attachment detected, but xlrd is not installed. Install xlrd to read .xls files with pandas."
        except Exception as exc:
            return f"XLS pandas read failed: {type(exc).__name__}: {exc}"

        sections: list[str] = []
        for sheet_name, dataframe in sheets.items():
            preview = dataframe.head(self.config.max_table_rows)
            rows = [list(map(str, preview.columns.tolist()))]
            for _, row in preview.iterrows():
                rows.append(["" if value is None else str(value) for value in row.tolist()])

            section_parts = [
                format_rows(
                    f"Sheet {sheet_name} ({len(dataframe)} rows x {len(dataframe.columns)} columns)",
                    rows,
                    truncated=len(dataframe) > self.config.max_table_rows,
                    max_rows=self.config.max_table_rows,
                )
            ]
            summary = self._summarize_excel_rows(rows)
            if summary:
                section_parts.append(summary)
            question_summary = self._build_question_oriented_excel_summary(question, rows)
            if question_summary:
                section_parts.append(question_summary)
            sections.append("\n\n".join(section_parts).strip())

        return "\n\n".join(section for section in sections if section)

    def _safe_int(self, value: Any, default: int = 0) -> int:
        """
        負責執行 ExcelAttachmentReader 中的 _safe_int 流程，依照 ExcelAttachmentReader 的流程需求處理 _safe_int 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            value: 此流程需要使用的輸入資料。
            default: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 int。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        try:
            return int(value or default)
        except Exception:
            return default

    def _get_excel_fill_color(self, cell: Any) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 _get_excel_fill_color 流程，依照 ExcelAttachmentReader 的流程需求處理 _get_excel_fill_color 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            cell: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        fill = getattr(cell, "fill", None)
        if fill is None or getattr(fill, "fill_type", None) is None:
            return ""

        color = getattr(fill, "fgColor", None)
        if color is None:
            return ""

        color_type = str(getattr(color, "type", "") or "").lower()
        if color_type == "rgb" and getattr(color, "rgb", None):
            return str(color.rgb)
        if color_type == "indexed" and getattr(color, "indexed", None) is not None:
            return f"indexed:{color.indexed}"
        if color_type == "theme" and getattr(color, "theme", None) is not None:
            tint = getattr(color, "tint", None)
            suffix = f":tint={tint}" if tint not in {None, 0, 0.0} else ""
            return f"theme:{color.theme}{suffix}"
        return ""

    def _format_color_rows(self, color_rows: list[list[str]]) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 _format_color_rows 流程，依照 ExcelAttachmentReader 的流程需求處理 _format_color_rows 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            color_rows: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        lines = ["Cell colors:"]
        for idx, row in enumerate(color_rows[: self.config.max_table_rows], start=1):
            lines.append(f"row {idx}: " + " | ".join(row))
        if len(color_rows) > self.config.max_table_rows:
            lines.append(f"[showing first {self.config.max_table_rows} color rows]")
        return "\n".join(lines)

    def _classify_excel_question(self, question: str) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 _classify_excel_question 流程，依照 ExcelAttachmentReader 的流程需求處理 _classify_excel_question 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            question: 目前要處理的任務、問題或查詢文字。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        normalized = str(question or "").lower()
        color_markers = [
            "color",
            "colour",
            "cell color",
            "green cell",
            "plot",
            "land",
            "map",
            "start",
            "end",
            "path",
            "walk",
            "move",
            "maze",
            "grid",
            "route",
        ]
        table_markers = [
            "sales",
            "total",
            "sum",
            "food",
            "drink",
            "price",
            "amount",
            "revenue",
            "count",
            "average",
        ]
        if any(marker in normalized for marker in color_markers):
            return "color_grid"
        if any(marker in normalized for marker in table_markers):
            return "table_summary"
        return "balanced"

    def _excel_row_limit_for_strategy(self, strategy: str) -> int:
        """
        負責執行 ExcelAttachmentReader 中的 _excel_row_limit_for_strategy 流程，依照 ExcelAttachmentReader 的流程需求處理 _excel_row_limit_for_strategy 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            strategy: 控制檢索、篩選或輸出數量的數值參數。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 int。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        if strategy == "color_grid":
            return self.config.max_table_rows
        if strategy == "table_summary":
            return min(self.config.max_table_rows, 40)
        return min(self.config.max_table_rows, 60)

    def _build_excel_color_section(
        self,
        *,
        strategy: str,
        color_rows: list[list[str]],
        colored_cells: list[str],
        max_row: int,
        max_column: int,
    ) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 _build_excel_color_section 流程，依照 ExcelAttachmentReader 的流程需求處理 _build_excel_color_section 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            strategy: 此流程需要使用的輸入資料。
            color_rows: 此流程需要使用的輸入資料。
            colored_cells: 此流程需要使用的輸入資料。
            max_row: 控制檢索、篩選或輸出數量的數值參數。
            max_column: 控制檢索、篩選或輸出數量的數值參數。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        if not color_rows and not colored_cells:
            return ""

        if strategy == "color_grid":
            return self._format_color_rows(color_rows)

        color_count = sum(len(row) for row in color_rows) if color_rows else len(colored_cells)
        unique_colors = self._unique_excel_colors(color_rows, colored_cells)
        lines = [
            "Cell color summary:",
            f"- sheet_size: {max_row} rows x {max_column} columns",
            f"- colored_cell_count_in_preview: {color_count}",
            f"- unique_colors_in_preview: {', '.join(unique_colors[:20]) if unique_colors else '(none)'}",
        ]
        if strategy == "balanced":
            lines.append("Sample colored cells:")
            lines.extend(f"- {cell}" for cell in colored_cells[:20])
        else:
            lines.append("- full color grid omitted because this question is table-oriented")
        return "\n".join(lines)

    def _unique_excel_colors(self, color_rows: list[list[str]], colored_cells: list[str]) -> list[str]:
        """
        負責執行 ExcelAttachmentReader 中的 _unique_excel_colors 流程，依照 ExcelAttachmentReader 的流程需求處理 _unique_excel_colors 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            color_rows: 此流程需要使用的輸入資料。
            colored_cells: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 list[str]。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        colors: set[str] = set()
        for row in color_rows:
            for item in row:
                if "=" in item:
                    colors.add(item.split("=", 1)[1].strip())
        for item in colored_cells:
            marker = " color="
            if marker in item:
                colors.add(item.split(marker, 1)[1].strip())
        return sorted(colors)

    def _summarize_excel_rows(self, rows: list[list[str]]) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 _summarize_excel_rows 流程，依照 ExcelAttachmentReader 的流程需求處理 _summarize_excel_rows 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            rows: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        if len(rows) < 2:
            return ""

        header = [
            compact_single_line(value, default=f"column_{idx + 1}")
            for idx, value in enumerate(rows[0])
        ]
        data = rows[1:]
        lines = ["Column summary:"]
        added = False
        for col_idx, name in enumerate(header):
            values = [row[col_idx] if col_idx < len(row) else "" for row in data]
            non_empty = [compact_single_line(value, default="") for value in values]
            non_empty = [value for value in non_empty if value]
            if not non_empty:
                continue

            numeric_values: list[float] = []
            text_values: list[str] = []
            for value in non_empty:
                parsed = self._try_parse_number(value)
                if parsed is None:
                    text_values.append(value)
                else:
                    numeric_values.append(parsed)

            if numeric_values and len(numeric_values) >= max(1, len(non_empty) // 2):
                lines.append(
                    f"- {name}: numeric_count={len(numeric_values)} "
                    f"sum={sum(numeric_values):.2f} min={min(numeric_values):.2f} max={max(numeric_values):.2f}"
                )
                added = True
            elif text_values:
                top_values = Counter(text_values).most_common(8)
                compact = ", ".join(f"{value}={count}" for value, count in top_values)
                lines.append(f"- {name}: top_values {compact}")
                added = True

        return "\n".join(lines) if added else ""

    def _build_question_oriented_excel_summary(self, question: str, rows: list[list[str]]) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 _build_question_oriented_excel_summary 流程，依照 ExcelAttachmentReader 的流程需求處理 _build_question_oriented_excel_summary 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            question: 目前要處理的任務、問題或查詢文字。
            rows: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        normalized_question = str(question or "").lower()
        if not rows or not ("food" in normalized_question and "drink" in normalized_question):
            return ""

        header = [
            compact_single_line(value, default=f"column_{idx + 1}")
            for idx, value in enumerate(rows[0])
        ]
        data = rows[1:]
        category_indices = [
            idx
            for idx, name in enumerate(header)
            if any(marker in name.lower() for marker in ["category", "type", "item", "menu"])
        ]
        numeric_indices = [
            idx
            for idx, name in enumerate(header)
            if any(marker in name.lower() for marker in ["sale", "sales", "total", "amount", "revenue", "price"])
        ]

        candidates: list[str] = []
        item_column_candidate = self._build_food_total_from_item_columns(header, data)
        if item_column_candidate:
            candidates.append(item_column_candidate)

        for category_idx in category_indices:
            for numeric_idx in numeric_indices:
                total = 0.0
                matched = 0
                excluded = 0
                for row in data:
                    category = compact_single_line(
                        row[category_idx] if category_idx < len(row) else "",
                        default="",
                    ).lower()
                    amount = self._try_parse_number(row[numeric_idx] if numeric_idx < len(row) else "")
                    if amount is None:
                        continue
                    if "drink" in category or "beverage" in category:
                        excluded += 1
                        continue
                    if category:
                        matched += 1
                        total += amount
                if matched:
                    candidates.append(
                        f"- sum {header[numeric_idx]} where {header[category_idx]} is not drink/beverage: "
                        f"{total:.2f} (included_rows={matched}, excluded_rows={excluded})"
                    )

        if not candidates:
            return ""
        return "Question-oriented calculations:\n" + "\n".join(candidates[:6])

    def _build_food_total_from_item_columns(self, header: list[str], data: list[list[str]]) -> str:
        """
        負責執行 ExcelAttachmentReader 中的 _build_food_total_from_item_columns 流程，依照 ExcelAttachmentReader 的流程需求處理 _build_food_total_from_item_columns 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            header: 此流程需要使用的輸入資料。
            data: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 str。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        drink_markers = {"drink", "drinks", "beverage", "beverages", "soda", "coffee", "tea", "juice", "water"}
        non_item_markers = {"location", "store", "branch", "city", "date", "id", "name"}
        included_columns: list[str] = []
        excluded_columns: list[str] = []
        total = 0.0

        for col_idx, name in enumerate(header):
            normalized_name = name.lower()
            if normalized_name in non_item_markers or any(marker in normalized_name for marker in non_item_markers):
                continue

            numeric_values: list[float] = []
            for row in data:
                amount = self._try_parse_number(row[col_idx] if col_idx < len(row) else "")
                if amount is not None:
                    numeric_values.append(amount)
            if not numeric_values:
                continue

            if normalized_name in drink_markers or any(marker in normalized_name for marker in drink_markers):
                excluded_columns.append(name)
                continue

            included_columns.append(name)
            total += sum(numeric_values)

        if not included_columns:
            return ""
        excluded_text = ", ".join(excluded_columns) if excluded_columns else "(none)"
        return (
            f"- sum numeric item columns excluding drink-like columns: {total:.2f} "
            f"(included_columns={', '.join(included_columns)}; excluded_columns={excluded_text})"
        )

    def _try_parse_number(self, value: Any) -> float | None:
        """
        負責執行 ExcelAttachmentReader 中的 _try_parse_number 流程，依照 ExcelAttachmentReader 的流程需求處理 _try_parse_number 對應的資料轉換、狀態操作或結果產生。
        
        Args:
            value: 此流程需要使用的輸入資料。
        
        Returns:
            執行結果；若函式標註回傳型別，預期型別為 float | None。
        
        限制或副作用:
            可能讀取或更新物件狀態、檔案、外部服務或日誌；請依呼叫場景確認副作用。
        """
        text = compact_single_line(value, default="")
        if not text:
            return None
        text = text.replace(",", "").replace("$", "").replace("%", "")
        try:
            return float(text)
        except Exception:
            return None
